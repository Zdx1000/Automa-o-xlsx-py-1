import openpyxl
import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import threading
import tkinter as tk
from tkinter import ttk
import os

#timer e geração processo
def gerar_arquivo():
    total_steps = 10
    current_step = 0

#processo animado
    def update_progress(step_increment=1, message=""):
        nonlocal current_step
        current_step += step_increment
        progress_percentage = int((current_step / total_steps) * 100)
        progress_bar['value'] = progress_percentage
        progress_label.config(text=f"{message} ({progress_percentage}%)")
        janela.update_idletasks()

    update_progress(message="Carregando dados...")

    #config manual PDS retirar
    excel_dados = pd.read_excel("Config.xlsx", sheet_name="Princial 1")
    pasta1 = excel_dados.loc[excel_dados["index"] == 1, "Endereços a retirar Auto.app [ Diretório ]"]
    retirar_end = str(pasta1.iloc[0]) #Endereço a retirar
    pasta2 = excel_dados.loc[excel_dados["index"] == 2, "Endereços a retirar Auto.app [ Diretório ]"]
    nome_end = str(pasta2.iloc[0]) #Nome do arquivo xlsx
    pasta3 = excel_dados.loc[excel_dados["index"] == 3, "Endereços a retirar Auto.app [ Diretório ]"]
    planilha_reti = str(pasta3.iloc[0]) #Nome planilha
    pasta9 = excel_dados.loc[excel_dados["index"] == 1, "Tipo dos arquivos Auto.app [ Tipos ]"]
    tipo_reti = str(pasta9.iloc[0]) #tipo retirada

    #config dados P&D
    pasta4 = excel_dados.loc[excel_dados["index"] == 1, "Endereços Auto.app [ Diretório alimentação ]"]
    diretorio = str(pasta4.iloc[0]) #Alimentação

    #Config dados modulos P&D
    pasta10 = excel_dados.loc[excel_dados["index"] == 6, "Endereços Auto.app [ Diretório alimentação ]"]
    ender_modulo = str(pasta10.iloc[0]) #Endereço Modulos
    pasta11 = excel_dados.loc[excel_dados["index"] == 7, "Endereços Auto.app [ Diretório alimentação ]"]
    nome_modulo = str(pasta11.iloc[0]) #Nomes modulos sem BK
    pasta12 = excel_dados.loc[excel_dados["index"] == 8, "Endereços Auto.app [ Diretório alimentação ]"]
    nome_modulo_BK = str(pasta12.iloc[0]) #Nomes modulos com BK
    pasta15 = excel_dados.loc[excel_dados["index"] == 9, "Endereços Auto.app [ Diretório alimentação ]"]
    nome_planilha_modulo = str(pasta15.iloc[0]) #Nomes modulos com BK

    #Config data P&D
    pasta16 = excel_dados.loc[excel_dados["index"] == 6, "Endereços a retirar Auto.app [ Diretório ]"]
    ender_PD = str(pasta16.iloc[0]) #Nomes modulos com BK
    pasta17 = excel_dados.loc[excel_dados["index"] == 7, "Endereços a retirar Auto.app [ Diretório ]"]
    nome_PD = str(pasta17.iloc[0]) #Nomes modulos com BK
    pasta18 = excel_dados.loc[excel_dados["index"] == 6, "Tipo dos arquivos Auto.app [ Tipos ]"]
    tipo_PD = str(pasta18.iloc[0]) #Nomes modulos com BK

    #condig export .XLSX
    pasta19 = excel_dados.loc[excel_dados["index"] == 12, "Endereços Auto.app [ Diretório alimentação ]"]
    export = str(pasta19.iloc[0]) #Nomes modulos com BK
    pasta20 = excel_dados.loc[excel_dados["index"] == 13, "Endereços Auto.app [ Diretório alimentação ]"]
    tipo_export = str(pasta20.iloc[0]) #Nomes modulos com BK
    pasta21 = excel_dados.loc[excel_dados["index"] == 14, "Endereços Auto.app [ Diretório alimentação ]"]
    nome_export = str(pasta21.iloc[0]) #Nomes modulos com BK


    #config Nome dos arquivos
    excel_dados_2 = pd.read_excel("Config.xlsx", sheet_name="Princial 2")
    pasta5 = excel_dados_2.loc[excel_dados_2["index"] == 1, "Nomes dos arquivos Auto.app [ Nomes ]"]
    arquivo1_nome = str(pasta5.iloc[0]) #Nome do arquivo 1
    pasta6 = excel_dados_2.loc[excel_dados_2["index"] == 2, "Nomes dos arquivos Auto.app [ Nomes ]"]
    arquivo2_nome = str(pasta6.iloc[0]) #Nome do arquivo 2

    #config tipos dos aquivos
    pasta7 = excel_dados_2.loc[excel_dados_2["index"] == 1, "Tipo dos arquivos Auto.app [ Tipos ]"]
    tipo1 = str(pasta7.iloc[0]) #Tipo 1
    pasta8 = excel_dados_2.loc[excel_dados_2["index"] == 2, "Tipo dos arquivos Auto.app [ Tipos ]"]
    tipo2 = str(pasta8.iloc[0]) #Tipo 2

    #configg tipos dos modulos
    pasta13 = excel_dados_2.loc[excel_dados_2["index"] == 1, "Tipo dos aqeuivos Auto.app [ Tipo Modulos]"]
    tipo_modulo1 = str(pasta13.iloc[0]) #Tipo 1 modulo
    pasta14 = excel_dados_2.loc[excel_dados_2["index"] == 2, "Tipo dos aqeuivos Auto.app [ Tipo Modulos]"]
    tipo_modulo2 = str(pasta14.iloc[0]) #Tipo 2 modulo



    lista_sicro = []
    for sicros in os.listdir(diretorio):
        if sicros.startswith(f'{arquivo1_nome}') and sicros.endswith(f'{tipo1}') or sicros.endswith(f'{tipo2}'):
            caminho_arquivo = os.path.join(diretorio, sicros)

            sic = pd.read_excel(caminho_arquivo)

            lista_sicro.append(sic)
    sincronismo = pd.concat(lista_sicro, ignore_index=True)

    lista_uni = []
    for unit1 in os.listdir(diretorio):
        if unit1.startswith(f'{arquivo2_nome}') and unit1.endswith(f'{tipo1}') or unit1.endswith(f'{tipo2}'):
            caminho_arquivo = os.path.join(diretorio, unit1)

            uni1 = pd.read_excel(caminho_arquivo)

            lista_uni.append(uni1)
    mascle_utilizador = pd.concat(lista_uni, axis=0, ignore_index=True)

    update_progress(step_increment=1, message="Dados carregados")

    update_progress(message="Processando dados de utilizadores...")



    Enderecos_para_retirar = pd.read_excel(rf"{retirar_end}\{nome_end}{tipo_reti}", sheet_name=f"{planilha_reti}")

    modulos = pd.read_excel(rf"{ender_modulo}\{nome_modulo}{tipo_modulo1}", sheet_name=f"{nome_planilha_modulo}")

    modulos1 = pd.read_excel(rf"{ender_modulo}\{nome_modulo_BK}{tipo_modulo2}", sheet_name=f"{nome_planilha_modulo}")


    list = ["Cod. Mercadoria", "Descrição"]
    sincronismo = sincronismo[list]

    sincronismo = sincronismo.drop_duplicates("Cod. Mercadoria")
    mascle_utilizador = mascle_utilizador[mascle_utilizador["Carga"].isnull()]
    mascle_utilizador["Modulo"] = mascle_utilizador["Endereço"].str[0]

    list_mascle = ["Unitizador", "Item", "Modulo", "Endereço", "Descrição", "Qtde", "Custo Unit.", "Custo Total"]
    mascle_utilizador = mascle_utilizador[list_mascle]
    update_progress(step_increment=1, message="Dados de utilizadores processados")

    update_progress(message="Filtrando endereços para retirada...")
    retirada_condcoes = pd.merge(mascle_utilizador, Enderecos_para_retirar, how="left", left_on="Endereço", right_on="Endereços_retirada")
    retirada_condcoes = retirada_condcoes[retirada_condcoes["Endereços_retirada"].isnull()]
    masclar_base = pd.merge(retirada_condcoes, sincronismo, left_on="Item", right_on="Cod. Mercadoria", how="inner")
    update_progress(step_increment=1, message="Endereços filtrados")

    update_progress(message="Preparando dados finais...")
    data_atual = datetime.now()
    data_hoje = data_atual.date()
    masclar_base["Data_Geração"] = data_hoje
    masclar_base["Dias no P&D"] = data_hoje - masclar_base["Data_Geração"]

    list2 = ["Data_Geração", "Unitizador", "Item", "Modulo", "Endereço", "Descrição_x", "Qtde", "Custo Unit.", "Custo Total", "Dias no P&D"]
    masclar_base = masclar_base[list2]
    masclar_modulos = pd.merge(masclar_base, modulos, how="left", left_on="Endereço", right_on="Endereço_MODULOS")
    masclar_modulos = pd.DataFrame(masclar_modulos)

    modulos_index = ["Data_Geração", "Unitizador", "Item", "Modulos", "Endereço", "Descrição_x", "Qtde", "Custo Unit.", "Custo Total", "Dias no P&D"]
    masclar_modulos = masclar_modulos[modulos_index]
    update_progress(step_increment=1, message="Dados finais preparados")

    update_progress(message="Salvando dados em arquivo...")
    d = {"index": [1],
         "Data": [f"PD_{data_hoje}"]}
    planilha_data = pd.DataFrame(data=d)

    masclar_modulos["Descrição do item"] = masclar_modulos["Descrição_x"]
    masclar_modulos = masclar_modulos.drop(columns=["Descrição_x"])

    try:
        lista_pd = []
        for PD in os.listdir(ender_PD):
            if PD.startswith('PD_') and (PD.endswith('xlsx') or PD.endswith('xls')):
                caminho_arquivo = os.path.join(ender_PD, PD)

                try:
                    sic = pd.read_excel(caminho_arquivo, sheet_name='Geral')
                    lista_pd.append(sic)
                except ValueError:
                    print(f'A planilha "Geral" não foi encontrada no arquivo: {PD}')
                except Exception as e:
                    print(f'Erro ao ler o arquivo {PD}: {e}')

        varificao = pd.concat(lista_pd, ignore_index=True)

        varificao['Data_Geração'] = pd.to_datetime(varificao['Data_Geração'], errors='coerce')

        varificao = varificao.sort_values(by=['Unitizador', 'Data_Geração'])

        varificao = varificao.drop_duplicates(subset='Unitizador', keep='first')

        list_v = ["Data_Geração", "Unitizador"]
        varificao = varificao[list_v]

        varificao['Unitizador'] = pd.to_numeric(varificao['Unitizador'], errors='coerce')
        masclar_modulos['Unitizador'] = pd.to_numeric(masclar_modulos['Unitizador'], errors='coerce')

        varificao['Unitizador'] = varificao['Unitizador'].astype(int)
        masclar_modulos['Unitizador'] = masclar_modulos['Unitizador'].astype(int)

        teste = pd.merge(varificao, masclar_modulos, left_on="Unitizador", right_on="Unitizador", how="right")
        teste['Data_Geração'] = teste['Data_Geração_x']

        if 'Data_Geração_y' in teste.columns:
            teste['Data_Geração'] = teste['Data_Geração'].combine_first(teste['Data_Geração_y'])
            teste = teste.drop(columns=['Data_Geração_y'])
            teste = teste.drop(columns=['Data_Geração_x'])

        lista_ori = ["Data_Geração", "Unitizador", "Item", "Modulos", "Endereço", "Descrição do item", "Qtde", "Custo Unit.", "Custo Total", "Dias no P&D"]
        teste = teste[lista_ori]

        teste['Data_Geração'] = pd.to_datetime(teste['Data_Geração']).dt.date

        teste['Dias no P&D'] = (data_hoje - teste['Data_Geração']).apply(lambda x: x.days)

        masclar_modulos = pd.DataFrame(teste)
    except:
        masclar_modulos = pd.DataFrame(masclar_modulos)

    modulos_unicos = masclar_modulos["Modulos"].unique()
    modulos_amare = ["Unitizador", "Item", "Endereço", "Descrição do item", "Qtde", "Modulos"]

    modulos_amare_b = masclar_modulos[modulos_amare]

    masclar_final = pd.merge(masclar_modulos, modulos1, how="left", left_on="Endereço", right_on="Endereço_MODULOS")
    masclar_final["Modulos"] = masclar_final["Modulos_y"]
    masclar_final = masclar_final.drop(columns=["Modulos_y"])
    masclar_final = masclar_final.drop(columns=["Modulos_x"])

    dinamica = masclar_final.pivot_table(index='Dias no P&D', columns='Modulos', values='Unitizador',
                                              aggfunc='count').reset_index()


    dinamica.loc['Total'] = dinamica.select_dtypes(include=[int, float]).sum()

    dinamica.at['Total', 'Dias no P&D'] = 'Total'

    dinamica['Dias no P&D'] = dinamica['Dias no P&D'].apply(
        lambda x: f"{int(x)} Dia no P&D" if isinstance(x, (int, float)) else x)

    masclar_final['Custo Total'] = masclar_final['Custo Total'].str.replace(',', '.').astype(float)

    modulos_vlr_cont = masclar_final.pivot_table(index='Modulos', values=["Custo Total", "Unitizador"],
                                                 aggfunc={ "Custo Total": "sum", "Unitizador" : "count"}).reset_index()

    modulos_vlr_cont.loc['Total'] = modulos_vlr_cont.select_dtypes(include=[int, float]).sum()

    modulos_vlr_cont.at['Total', 'Modulos'] = 'Total'


    planilha_data.to_excel(rf"{ender_PD}\{nome_PD}{tipo_PD}", index=False)
    arquivo_base = rf"{export}\{nome_export}{data_hoje}{tipo_export}"
    with pd.ExcelWriter(arquivo_base, engine="openpyxl") as writer:
        dinamica.to_excel(writer, sheet_name="Dias no P&D", index=False)
        modulos_vlr_cont.to_excel(writer, sheet_name="Modulos", index=False)
        masclar_final.to_excel(writer, sheet_name="Geral", index=False)
        varificao.to_excel(writer, sheet_name="teste", index=False)
        for modulo in modulos_unicos:
            df_modulo = modulos_amare_b[modulos_amare_b["Modulos"] == modulo]
            df_modulo.to_excel(writer, sheet_name=f"{modulo}", index=False)

    workbook = openpyxl.load_workbook(arquivo_base)
    worksheet = workbook["Dias no P&D"]

    header_fill = PatternFill(start_color="483D8B", end_color="483D8B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for col_num in range(1, len(dinamica.columns) + 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet[f"{col_letter}1"]
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = header_font

        cell_total = worksheet[f"{col_letter}{worksheet.max_row}"]
        cell_total.fill = total_fill
        cell_total.border = thin_border
        cell_total.alignment = Alignment(horizontal="center", vertical="center")
        cell_total.font = Font(bold=True)

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row - 1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column].width = adjusted_width

    worksheet = workbook["Modulos"]

    for col_num in range(1, len(modulos_vlr_cont.columns) + 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet[f"{col_letter}1"]
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = header_font

        cell_total = worksheet[f"{col_letter}{worksheet.max_row}"]
        cell_total.fill = total_fill
        cell_total.border = thin_border
        cell_total.alignment = Alignment(horizontal="center", vertical="center")
        cell_total.font = Font(bold=True)

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row - 1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column].width = adjusted_width

    worksheet = workbook["Geral"]

    for col_num in range(1, len(masclar_final.columns) + 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet[f"{col_letter}1"]
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = header_font

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column].width = adjusted_width


    for ed in modulos_unicos:
        worksheet = workbook[f"{ed}"]

        for col_num in range(1, len(modulos_amare_b.columns) + 1):
            col_letter = get_column_letter(col_num)
            cell = worksheet[f"{col_letter}1"]
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = header_font

            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(arquivo_base)
    workbook.close()
    update_progress(step_increment=1, message="Arquivo salvo com sucesso")

    iniciar_button.config(text="Fechar", state=tk.NORMAL, command=janela.destroy)

def iniciar_processo():
    iniciar_button.config(text="Gerando...", state=tk.DISABLED)
    threading.Thread(target=gerar_arquivo).start()

janela = tk.Tk()
janela.title("Geração de Arquivo")
janela.geometry("600x300")
janela.configure(bg="#2C3E50")
janela.resizable(False, False)
janela.iconbitmap(r"icone.ico")

title_label = tk.Label(janela, text="Processo de Geração de Arquivo", font=('Helvetica', 20, 'bold'), fg="white", bg="#2C3E50")
title_label.pack(pady=20)

progress_label = tk.Label(janela, text="Aguardando Início", font=('Helvetica', 14), fg="white", bg="#2C3E50")
progress_label.pack(pady=10)

progress_frame = tk.Frame(janela, bg="#2C3E50")
progress_frame.pack(pady=10)

progress_bar_style = ttk.Style()
progress_bar_style.theme_use('clam')
progress_bar_style.configure("Horizontal.TProgressbar", troughcolor='#34495E', background='#1E90FF', thickness=30)

progress_bar = ttk.Progressbar(progress_frame, orient="horizontal", length=500, mode="determinate", style="Horizontal.TProgressbar")
progress_bar.pack(side=tk.LEFT, padx=10)

progress_percent = tk.Label(progress_frame, text="0%", font=('Helvetica', 12), fg="white", bg="#2C3E50")
progress_percent.pack(side=tk.LEFT)

def update_progress(value):
    progress_bar['value'] = value
    progress_percent.config(text=f"{value}%")
    janela.update_idletasks()

iniciar_button = tk.Button(janela, text="Iniciar Geração", command=iniciar_processo, font=('Helvetica', 14), bg="#1E90FF", fg="white", cursor="hand2")
iniciar_button.pack(pady=20)

janela.mainloop()
